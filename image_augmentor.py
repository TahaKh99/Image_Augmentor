import os
import random
import numpy as np
from PIL import Image
from keras.preprocessing.image import ImageDataGenerator
from skimage import io
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QLabel,
    QLineEdit,
    QPushButton,
    QCheckBox,
    QVBoxLayout,
    QWidget,
    QFileDialog,
    QMessageBox,
    QHBoxLayout,
    QComboBox
)
from PyQt5.QtGui import QFont, QIcon, QPalette, QColor
from PyQt5.QtCore import Qt

from imgaug import augmenters as iaa
import albumentations as A
import torchvision.transforms as transforms
import torch
import torchvision
import xml.etree.ElementTree as ET
import openpyxl
class ImageAugmentor(QMainWindow):
    @staticmethod
    def add_noise(img):
        """Add random noise to an image"""
        VARIABILITY = 50
        deviation = VARIABILITY * random.random()
        noise = np.random.normal(0, deviation, img.shape)
        img = np.clip(img + noise, 0, 255).astype(np.uint8)
        return img
    def add_noise_with_imgaug(self, img):
        seq = iaa.SaltAndPepper(0.05)
        noisy_img = seq.augment_image(img)
        return noisy_img

    def add_noise_with_albumentations(self, img):
        transform = A.Compose([A.OneOf([
            A.GaussNoise(p=1),
            A.MultiplicativeNoise(p=1),
        ], p=1)])
        noisy_img = transform(image=img)["image"]
        return noisy_img

    def add_noise_with_torchvision(self, img):
        transform = torchvision.transforms.ToTensor()
        img_tensor = transform(img)
        noise_transform = torchvision.transforms.Compose([
            torchvision.transforms.ToPILImage(),
            torchvision.transforms.ToTensor(),
        ])
        noise = torch.randn_like(img_tensor)    
        noisy_img = img_tensor + noise
        noisy_img = noise_transform(noisy_img)
        return noisy_img
    def generate_annotation_file(self):
        try:
            image_directory = self.image_directory_edit.text()
            annotations_directory = self.save_directory_edit.text()

            for filename in os.listdir(image_directory):
                image_path = os.path.join(image_directory, filename)
                if os.path.isfile(image_path):
                    annotation_path = os.path.join(annotations_directory, f"{os.path.splitext(filename)[0]}.xml")
                    self.generate_voc_annotation(image_path, annotation_path)

            QMessageBox.information(self,"Success", "Annotation files generated successfully.")
        except Exception as e:
            QMessageBox.warning(self,"Error", str(e))

    def generate_voc_annotation(self, image_path, annotation_path):
        try:
            # Extract image dimensions
            width, height = self.get_image_dimensions(image_path)

            # Define the bounding box coordinates
            xmin, ymin, xmax, ymax = 100, 100, 200, 200  # Example values

            # Create the XML annotation tree
            root = ET.Element("annotation")
            folder = ET.SubElement(root, "folder")
            folder.text = os.path.dirname(image_path)
            filename = ET.SubElement(root, "filename")
            filename.text = os.path.basename(image_path)
            size = ET.SubElement(root, "size")
            width_elem = ET.SubElement(size, "width")
            width_elem.text = str(width)
            height_elem = ET.SubElement(size, "height")
            height_elem.text = str(height)
            depth = ET.SubElement(size, "depth")
            depth.text = "3"  # Assuming RGB images

            object_elem = ET.SubElement(root, "object")
            name = ET.SubElement(object_elem, "name")
            name.text = "object_class"  # Replace with the actual class label
            bndbox = ET.SubElement(object_elem, "bndbox")
            xmin_elem = ET.SubElement(bndbox, "xmin")
            xmin_elem.text = str(xmin)
            ymin_elem = ET.SubElement(bndbox, "ymin")
            ymin_elem.text = str(ymin)
            xmax_elem = ET.SubElement(bndbox, "xmax")
            xmax_elem.text = str(xmax)
            ymax_elem = ET.SubElement(bndbox, "ymax")
            ymax_elem.text = str(ymax)

            # Create the annotation file
            tree = ET.ElementTree(root)
            tree.write(annotation_path)
        except Exception as e:
            raise Exception(self,"Error generating VOC annotation: " + str(e))

    def get_image_dimensions(self, image_path):
        try:
            # Implement a function to get the image dimensions (width, height)
            from PIL import Image
            image = Image.open(image_path)
            width, height = image.size
            return width, height
        except Exception as e:
            raise Exception("Error getting image dimensions: " + str(e))
        
    def generate_labeling_file(self):
        try:
            image_directory = self.image_directory_edit.text()
            labeling_file_directory = self.save_directory_edit.text()

            # Create the labeling file path
            labeling_file_path = os.path.join(labeling_file_directory, "labeling_file.xlsx")

            # Get the list of image files in the directory
            image_files = os.listdir(image_directory)

            # Create a new workbook
            workbook = openpyxl.Workbook()

            # Get the active sheet
            sheet = workbook.active

            # Write the header row
            sheet['A1'] = 'Image File'
            sheet['B1'] = 'Label'

            # Iterate over each image file
            for row, file in enumerate(image_files, start=2):
                # Extract the label from the filename or provide a default label
                label = self.extract_label_from_filename(file)

                # Write the image file and label in the corresponding cells
                sheet.cell(row=row, column=1).value = file
                sheet.cell(row=row, column=2).value = label

            # Save the workbook
            workbook.save(labeling_file_path)

            QMessageBox.information(self,"Success", "Labeling file generated successfully.")
        except Exception as e:
            QMessageBox.warning(self,"Error", str(e))

    # Function to extract the label from the filename
    def extract_label_from_filename(self, filename):
        # For example, if the filename is "cat001.jpg", you can extract the label as "cat"
        label = filename.split('.')[0]  # Assumes the label is before the first dot
        return label


    
    

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Image Augmentation")
        self.setWindowIcon(QIcon("E:\Final_Project\puppy.jpg"))
        self.setGeometry(100, 100, 800, 600)

        # Central Widget
        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)
        self.layout.setAlignment(Qt.AlignTop)
        self.layout.setContentsMargins(30, 30, 30, 30)

        # Title Label
        self.title_label = QLabel("Image Augmentation", self.central_widget)
        self.title_label.setStyleSheet(
            "font-size: 48px; font-weight: bold; color: #333333; margin-bottom: 30px;"
        )
        self.layout.addWidget(self.title_label, alignment=Qt.AlignCenter)

        # Image and Save Directory Section
        directory_layout = QHBoxLayout()
        self.image_directory_label = QLabel("Image Directory:", self.central_widget)
        self.save_directory_label = QLabel("Save Directory:", self.central_widget)
        directory_layout.addWidget(self.image_directory_label)
        directory_layout.addWidget(self.save_directory_label)
        self.layout.addLayout(directory_layout)

        directory_input_layout = QHBoxLayout()
        self.image_directory_edit = QLineEdit(self.central_widget)
        self.image_directory_edit.setReadOnly(True)
        self.save_directory_edit = QLineEdit(self.central_widget)
        self.save_directory_edit.setReadOnly(True)
        directory_input_layout.addWidget(self.image_directory_edit)
        directory_input_layout.addWidget(self.save_directory_edit)
        self.layout.addLayout(directory_input_layout)

        browse_layout = QHBoxLayout()
        self.select_image_directory_button = QPushButton("Browse", self.central_widget)
        self.select_image_directory_button.setObjectName("browse_button")
        self.select_image_directory_button.clicked.connect(self.select_image_directory)
        self.select_save_directory_button = QPushButton("Browse", self.central_widget)
        self.select_save_directory_button.setObjectName("browse_button")
        self.select_save_directory_button.clicked.connect(self.select_save_directory)
        self.select_save_directory_button.setStyleSheet(
            """
            QPushButton {
                font-size: 16px;
                font-weight: bold;
                color: #ffffff;
                background-color: #00a8e8;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #0091c9;
            }
            """
        )
        self.select_image_directory_button.setStyleSheet(
            """
            QPushButton {
                font-size: 16px;
                font-weight: bold;
                color: #ffffff;
                background-color: #00a8e8;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #0091c9;
            }
            """
        )
        browse_layout.addWidget(self.select_image_directory_button)
        browse_layout.addWidget(self.select_save_directory_button)
        self.layout.addLayout(browse_layout)

        # Augmentation Section
        self.augmentation_label = QLabel("Select the augmentation library:", self.central_widget)
        self.augmentation_combo_box = QComboBox(self.central_widget)
        self.augmentation_combo_box.addItem("ImageDataGenerator (Keras)")
        self.augmentation_combo_box.addItem("imgaug")
        self.augmentation_combo_box.addItem("albumentations")
        self.augmentation_combo_box.addItem("torchvision")

        self.layout.addWidget(self.augmentation_label)
        self.layout.addWidget(self.augmentation_combo_box)
        
        # Apply Augmentation Button
        self.apply_button = QPushButton("Apply Augmentation", self.central_widget)
        self.apply_button.setObjectName("apply_button")
        self.apply_button.clicked.connect(self.apply_augmentation)
        self.layout.addWidget(self.apply_button, alignment=Qt.AlignCenter)
        self.apply_button.setStyleSheet(
            """
            QPushButton {
                font-size: 20px;
                font-weight: bold;
                color: #ffffff;
                background-color: #00a8e8;
                border: none;
                padding: 15px 40px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #0091c9;
            }
            """
        )
        # Augmentation Checkboxes
        self.augmentation_checkboxes = []
        self.add_augmentation_checkbox("Rotate (7)")
        self.add_augmentation_checkbox("Shift (4)")
        self.add_augmentation_checkbox("Shear (3)")
        self.add_augmentation_checkbox("Zoom (4)")
        self.add_augmentation_checkbox("Flip (4)")
        self.add_augmentation_checkbox("Change Brightness (5)")
        self.add_augmentation_checkbox("Add Noise (5)")

        # Select All Checkbox
        self.select_all_checkbox = QCheckBox("Select All", self.central_widget)
        self.select_all_checkbox.stateChanged.connect(self.toggle_all)
        self.layout.addWidget(self.select_all_checkbox)

        # Generate Annotation and Labeling Buttons
        button_layout = QHBoxLayout()
        self.generate_annotation_button = QPushButton("Generate Annotation File", self.central_widget)
        self.generate_annotation_button.setObjectName("action_button")
        self.generate_annotation_button.clicked.connect(self.generate_annotation_file)
        self.generate_annotation_button.setStyleSheet(
            """
            QPushButton {
                font-size: 16px;
                font-weight: bold;
                color: #ffffff;
                background-color: #00a8e8;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #0091c9;
            }
            """
        )
        self.generate_labeling_button = QPushButton("Generate Labeling File", self.central_widget)
        self.generate_labeling_button.setObjectName("action_button")
        self.generate_labeling_button.clicked.connect(self.generate_labeling_file)
        self.generate_labeling_button.setStyleSheet(
            """
            QPushButton {
                font-size: 16px;
                font-weight: bold;
                color: #ffffff;
                background-color: #00a8e8;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #0091c9;
            }
            """
        )
        button_layout.addWidget(self.generate_annotation_button)
        button_layout.addWidget(self.generate_labeling_button)
        self.layout.addLayout(button_layout)

        # Apply the custom palette to the window
        self.set_palette()

    def add_augmentation_checkbox(self, label):
        checkbox = QCheckBox(label, self.central_widget)
        self.augmentation_checkboxes.append(checkbox)
        self.layout.addWidget(checkbox)

    def toggle_all(self, state):
        for checkbox in self.augmentation_checkboxes:
            checkbox.setChecked(state == Qt.Checked)

    def set_palette(self):
        # Define a custom color palette
        palette = QPalette()
        palette.setColor(QPalette.Background, QColor("#f9f9f9"))
        palette.setColor(QPalette.Base, QColor("#ffffff"))
        palette.setColor(QPalette.Text, QColor("#333333"))

        # Set the palette to the application
        self.setPalette(palette)


    def select_image_directory(self):
        image_directory = QFileDialog.getExistingDirectory(self, "Select Image Directory")
        self.image_directory_edit.setText(image_directory)

    def select_save_directory(self):
        save_directory = QFileDialog.getExistingDirectory(self, "Select Save Directory")
        self.save_directory_edit.setText(save_directory)

    def apply_augmentation(self):
        try:
            augmentation_folders = {
                    "Rotate (7)": "Rotated_Images",
                    "Shift (4)": "Shifted_Images",
                    "Shear (3)": "Sheared_Images",
                    "Zoom (4)": "Zoomed_Images",
                    "Flip (4)": "Flipped_Images",
                    "Change Brightness (5)": "Brightness_Images",
                    "Add Noise (5)": "Noisy_Images"
                                    }
            image_directory = self.image_directory_edit.text()
            save_directory = self.save_directory_edit.text()
            augmentation_library = self.augmentation_combo_box.currentText()

            for checkbox in self.augmentation_checkboxes:
                if checkbox.isChecked():
                    text = checkbox.text()
                    augmentation_folder = augmentation_folders.get(text, "Other")

                    # Create the folder for the augmentation method
                    augmentation_path = os.path.join(save_directory, augmentation_folder)
                    os.makedirs(augmentation_path, exist_ok=True)

                    if text.startswith("Rotate"):
                        if augmentation_library == "imgaug":
                            self.rotate_images_with_imgaug(image_directory,  num_images=7, augmentation_path = augmentation_path)
                        elif augmentation_library == "albumentations":
                            self.rotate_images_with_albumentations(image_directory,  num_images=7,augmentation_path = augmentation_path)
                        elif augmentation_library == "torchvision":
                            self.rotate_images_with_torchvision(image_directory,  num_images=7,augmentation_path = augmentation_path)
                        else:
                            self.rotate_images(image_directory, save_directory, num_images=7)

                    elif text.startswith("Shift"):
                        if augmentation_library == "imgaug":
                            self.shift_images_with_imagau(image_directory, num_images=4, augmentation_path = augmentation_path)
                        elif augmentation_library == "albumentations":
                            self.shift_images_with_albumentations(image_directory,num_images=4, augmentation_path = augmentation_path)
                        elif augmentation_library == "torchvision":
                            self.shift_images_with_torchvision(image_directory,  num_images=4, augmentation_path = augmentation_path)

                        else:
                            self.shift_images(image_directory, save_directory, num_images=4, )


                    elif text.startswith("Shear"):
                        if augmentation_library == "imgaug":
                            self.shear_images_with_imagau(image_directory,  num_images=4, augmentation_path = augmentation_path)
                        elif augmentation_library == "albumentations":
                            self.shear_images_with_albumentations(image_directory,  num_images=4, augmentation_path = augmentation_path)
                        elif augmentation_library == "torchvision":
                            self.shear_images_with_torchvision(image_directory,  num_images=4, augmentation_path = augmentation_path)

                        else:
                            self.shear_images(image_directory, save_directory, num_images=3)




                    elif text.startswith("Zoom"):
                        if augmentation_library == "imgaug":
                            self.zoom_images_with_imagau(image_directory, num_images=4, augmentation_path = augmentation_path)
                        elif augmentation_library == "albumentations":
                            self.zoom_images_with_albumentations(image_directory,  num_images=4, augmentation_path = augmentation_path)
                        elif augmentation_library == "torchvision":
                            self.zoom_images_with_torchvision(image_directory, num_images=4, augmentation_path = augmentation_path)

                        else:
                            self.zoom_images(image_directory, save_directory, num_images=4)



                    elif text.startswith("Flip"):
                        if augmentation_library == "imgaug":
                            self.flip_images_with_imagau(image_directory,  num_images=4, augmentation_path = augmentation_path)
                        elif augmentation_library == "albumentations":
                            self.flip_images_with_albumentations(image_directory,  num_images=4, augmentation_path = augmentation_path)
                        elif augmentation_library == "torchvision":
                            self.flip_images_with_torchvision(image_directory,  num_images=4, augmentation_path = augmentation_path)

                        else:
                            self.flip_images(image_directory, save_directory, num_images=4)



                    elif text.startswith("Change Brightness"):
                        if augmentation_library == "imgaug":
                            self.change_brightness_images_with_imagau(image_directory,  num_images=4, augmentation_path = augmentation_path)
                        elif augmentation_library == "albumentations":
                            self.change_brightness_images_with_albumentations(image_directory,  num_images=4, augmentation_path = augmentation_path)
                        elif augmentation_library == "torchvision":
                            self.change_brightness_images_with_torchvision(image_directory,  num_images=4, augmentation_path = augmentation_path)

                        else:
                            self.change_brightness(image_directory, save_directory, num_images=5)

                    elif text.startswith("Add Noise"):
                        if augmentation_library == "imgaug":
                            self.add_salt_pepper_noise_with_imagau(image_directory,  num_images=4, augmentation_path = augmentation_path)
                        elif augmentation_library == "albumentations":
                            self.add_salt_pepper_noise_with_albumentations(image_directory,  num_images=4, augmentation_path = augmentation_path)
                        elif augmentation_library == "torchvision":
                            self.add_salt_pepper_noise_images_with_torchvision()
                        else:
                            self.add_salt_pepper_noise(image_directory,  num_images=5,  augmentation_path = augmentation_path)

        except Exception as e:
            QMessageBox.warning(self,"Error", str(e))   

    def rotate_images(self, image_directory, save_directory, num_images):
            datagen = ImageDataGenerator(rotation_range=180)
            self.augment_images(datagen, image_directory, save_directory, "Rotated_img", num_images)
       


    def rotate_images_with_imgaug(self, image_directory, num_images, augmentation_path):
        # Create the augmentation sequence
        seq = iaa.Affine(rotate=(-180, 180))

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            img = Image.open(img_path)
            img_arr = np.array(img) 

            # Apply augmentation multiple times
            for i in range(num_images):
                augmented_img_arr = seq.augment_image(img_arr)
                augmented_img = Image.fromarray(augmented_img_arr)

                # Save the augmented image
                #save_path = os.path.join(save_directory, f"Rotated_{i + 1}_{file}")
                #augmented_img.save(save_path)
                new_file = f"Rotated_{i}_{file}"
                save_path = os.path.join(augmentation_path, new_file)
                augmented_img.save(save_path)
        QMessageBox.information(self, "Augmentation completed", "Image rotation augmentation completed!")

    def rotate_images_with_albumentations(self, image_directory, num_images, augmentation_path):
        # Create the augmentation transform
        transform = A.Rotate(limit=(-180, 180), p=1.0)

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            img = Image.open(img_path)

            # Apply augmentation multiple times
            for i in range(num_images):
                augmented = transform(image=np.array(img))
                augmented_img = Image.fromarray(augmented["image"]) 

                # Save the augmented image
                #save_path = os.path.join(save_directory, f"Rotated_{i + 1}_{file}")
                #augmented_img.save(save_path)
                new_file = f"Rotated_{i}_{file}"
                save_path = os.path.join(augmentation_path, new_file)
                augmented_img.save(save_path)

        QMessageBox.information(self, "Augmentation completed", "Image rotation augmentation completed!")

    def rotate_images_with_torchvision(self, image_directory, num_images, augmentation_path):
        # Create the rotation transform
        transform = transforms.RandomRotation(degrees=(-180, 180))

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            if os.path.isfile(img_path):
                # Load the image
                img = Image.open(img_path)

                # Apply augmentation to generate new images
                augmented_images = [transform(img) for _ in range(num_images)]

                # Save the augmented images
                for i, augmented_img in enumerate(augmented_images):
                    new_file = f"Rotated_{i}_{file}"
                    save_path = os.path.join(augmentation_path, new_file)
                    augmented_img.save(save_path)

        QMessageBox.information(self, "Augmentation completed", "Image rotation augmentation completed!")


#Shift .....................................

    def shift_images(self, image_directory, save_directory, num_images):
        datagen = ImageDataGenerator(width_shift_range=0.2, height_shift_range=0.2, fill_mode="reflect")
        self.augment_images(datagen, image_directory, save_directory, "Shifted_img", num_images)
    
    def shift_images_with_imagau(self, image_directory, num_images, augmentation_path):
        shift_range = (-120, 120)
        seq = iaa.Sequential([
            iaa.Affine(translate_px={"x": shift_range, "y": shift_range}, mode= 'reflect')
        ])

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            img = Image.open(img_path)
            img_arr = np.array(img)

            # Apply augmentation multiple times
            for i in range(num_images):
                augmented_img_arr = seq.augment_image(img_arr)
                augmented_img = Image.fromarray(augmented_img_arr)

                # Save the augmented image
                #save_path = os.path.join(save_directory, f"Rotated_{i + 1}_{file}")
                #augmented_img.save(save_path)
                new_file = f"Shifted{i}_{file}"
                save_path = os.path.join(augmentation_path, new_file)
                augmented_img.save(save_path)
        QMessageBox.information(self, "Augmentation completed", "Shift augmentation completed!")


    def shift_images_with_albumentations(self, image_directory, num_images, augmentation_path):
        # Create the augmentation transform
        transform = A.ShiftScaleRotate(shift_limit=0.2, p=1.0)

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            img = Image.open(img_path)
            img_arr = np.array(img)

            # Apply augmentation multiple times
            for i in range(num_images):
                augmented = transform(image=img_arr)
                augmented_img_arr = augmented["image"]
                augmented_img = Image.fromarray(augmented_img_arr)

                # Save the augmented image
                new_file = f"Shifted_{i}_{file}"
                save_path = os.path.join(augmentation_path, new_file)
                augmented_img.save(save_path)

        
        QMessageBox.information(self, "Augmentation completed", "Shift augmentation completed!")


    def shift_images_with_torchvision(self, image_directory, num_images, augmentation_path):
        # Create the augmentation transform
        transform = transforms.RandomAffine(degrees=0, translate=(0.2, 0.2), scale=None, shear=None)

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            if os.path.isfile(img_path):
                # Load the image
                img = Image.open(img_path)

                # Apply augmentation to generate new images
                augmented_images = [transform(img) for _ in range(num_images)]

                # Save the augmented images
                for i, augmented_img in enumerate(augmented_images):
                    new_file = f"Shifted_{i}_{file}"
                    save_path = os.path.join(augmentation_path, new_file)
                    augmented_img.save(save_path)


        QMessageBox.information(self, "Augmentation completed", "Shift augmentation completed!")

# Shear ................................................

    def shear_images(self, image_directory, save_directory, num_images):
        datagen = ImageDataGenerator(shear_range=45, fill_mode="reflect")
        self.augment_images(datagen, image_directory, save_directory, "Sheared_img", num_images)
        

    def shear_images_with_imagau(self, image_directory, num_images, augmentation_path):
        seq = iaa.Sequential([
        iaa.ShearX((-20, 20)),  # Apply shear transformation along the X-axis
        iaa.ShearY((-20, 20))   # Apply shear transformation along the Y-axis
    ])

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            img = Image.open(img_path)
            img_arr = np.array(img)

            # Apply augmentation multiple times
            for i in range(num_images):
                augmented_img_arr = seq.augment_image(img_arr)
                augmented_img = Image.fromarray(augmented_img_arr)

                # Save the augmented image
                new_file = f"Sheared_{i}_{file}"
                save_path = os.path.join(augmentation_path, new_file)
                augmented_img.save(save_path)
        QMessageBox.information(self, "Augmentation completed", "Shear augmentation completed!")


    def shear_images_with_albumentations(self, image_directory, num_images, augmentation_path):
        # Create the augmentation transform
        shear_angle = 20
        transform = A.ElasticTransform(p=1.0, alpha=shear_angle, sigma=10)

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            img = Image.open(img_path)
            img_arr = np.array(img)

            # Apply augmentation multiple times
            for i in range(num_images):
                augmented = transform(image=img_arr)
                augmented_img_arr = augmented["image"]
                augmented_img = Image.fromarray(augmented_img_arr)

                # Save the augmented image
                new_file = f"Sheared_{i}_{file}"
                save_path = os.path.join(augmentation_path, new_file)
                augmented_img.save(save_path)
    
        QMessageBox.information(self, "Augmentation completed", "Shear augmentation completed!")


    def shear_images_with_torchvision(self, image_directory,  num_images, augmentation_path):
        # Create the augmentation transform
        shear_angle = 20
        transform = transforms.RandomAffine(0, shear=(0, shear_angle))

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            if os.path.isfile(img_path):
                # Load the image
                img = Image.open(img_path)

                # Apply augmentation to generate new images
                augmented_images = [transform(img) for _ in range(num_images)]

                # Save the augmented images
                for i, augmented_img in enumerate(augmented_images):
                    new_file = f"Sheared_{i}_{file}"
                    save_path = os.path.join(augmentation_path, new_file)
                    augmented_img.save(save_path)

        QMessageBox.information(self, "Augmentation completed", "Shear augmentation completed!")




#Zoom .......................................

    def zoom_images(self, image_directory, save_directory, num_images):
            datagen = ImageDataGenerator(zoom_range=0.5, fill_mode="reflect")
            self.augment_images(datagen, image_directory, save_directory, "Zoomed_img", num_images)
        


    def zoom_images_with_imagau(self, image_directory, num_images, augmentation_path):
        seq = iaa.Affine(scale=(0.5, 1.5))

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            img = Image.open(img_path)
            img_arr = np.array(img)

            # Apply augmentation multiple times
            for i in range(num_images):
                augmented_img_arr = seq.augment_image(img_arr)
                augmented_img = Image.fromarray(augmented_img_arr)

                # Save the augmented image
                new_file = f"Zoomed_{i}_{file}"
                save_path = os.path.join(augmentation_path, new_file)
                augmented_img.save(save_path)

        QMessageBox.information(self, "Augmentation completed", "Zoom augmentation completed!")

    def zoom_images_with_albumentations(self, image_directory, num_images, augmentation_path):
        # Create the augmentation transform
        transform = A.RandomResizedCrop(height=512, width=512)

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            img = Image.open(img_path)
            img_arr = np.array(img)

            # Apply augmentation multiple times
            for i in range(num_images):
                augmented = transform(image=img_arr)
                augmented_img_arr = augmented["image"]
                augmented_img = Image.fromarray(augmented_img_arr)

                # Save the augmented image
                new_file = f"Zoomed_{i}_{file}"
                save_path = os.path.join(augmentation_path, new_file)
                augmented_img.save(save_path)
        
        QMessageBox.information(self, "Augmentation completed", "Zoom augmentation completed!")


    def zoom_images_with_torchvision(self, image_directory, num_images, augmentation_path):
        # Create the augmentation transform
        transform = transforms.RandomResizedCrop(size=(224, 224))

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            if os.path.isfile(img_path):
                # Load the image
                img = Image.open(img_path)

                # Apply augmentation to generate new images
                augmented_images = [transform(img) for _ in range(num_images)]

                # Save the augmented images
            for i, augmented_img in enumerate(augmented_images):
                new_file = f"Zoomed_{i}_{file}"
                save_path = os.path.join(augmentation_path, new_file)
                augmented_img.save(save_path)

        QMessageBox.information(self, "Augmentation completed", "Zoom augmentation completed!")


#Flip...................................................

    def flip_images(self, image_directory, save_directory, num_images):
        datagen = ImageDataGenerator(horizontal_flip=True, vertical_flip=True, fill_mode="reflect")
        self.augment_images(datagen, image_directory, save_directory, "Flipped_img", num_images)
   

    def flip_images_with_imagau(self, image_directory, num_images, augmentation_path):
        seq = iaa.Sequential([
        iaa.Fliplr(0.5),  # Apply horizontal flip with 50% probability
        iaa.Flipud(0.5)   # Apply vertical flip with 50% probability
        ])

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            img = Image.open(img_path)
            img_arr = np.array(img)

            # Apply augmentation multiple times
            for i in range(num_images):
                augmented_img_arr = seq.augment_image(img_arr)
                augmented_img = Image.fromarray(augmented_img_arr)

                # Save the augmented image
                new_file = f"flipped_Img{i}_{file}"
                save_path = os.path.join(augmentation_path, new_file)
                augmented_img.save(save_path)

        QMessageBox.information(self, "Augmentation completed", "Flip augmentation completed!")



    def flip_images_with_albumentations(self, image_directory, num_images, augmentation_path):
        # Create the augmentation transform
        transform = A.Compose([
                A.Flip(p=1.0),  # Flip the image horizontally or vertically with a probability of 1.0
            ])

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            img = Image.open(img_path)
            img_arr = np.array(img)

            # Apply augmentation multiple times
            for i in range(num_images):
                augmented = transform(image=img_arr)
                augmented_img_arr = augmented["image"]
                augmented_img = Image.fromarray(augmented_img_arr)

                # Save the augmented image
                new_file = f"flipped_Img{i}_{file}"
                save_path = os.path.join(augmentation_path, new_file)
                augmented_img.save(save_path)
        
        QMessageBox.information(self, "Augmentation completed", "Flip augmentation completed!")


    def flip_images_with_torchvision(self, image_directory,  num_images, augmentation_path):
        # Create the augmentation transform
        flip_transform = transforms.RandomChoice([
            transforms.RandomHorizontalFlip(p=1.0),
            transforms.RandomVerticalFlip(p=1.0)
        ])

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            if os.path.isfile(img_path):
                # Load the image
                img = Image.open(img_path)

                # Apply augmentation to generate new images
                augmented_images = [flip_transform(img) for _ in range(num_images)]

                # Save the augmented images
                for i, augmented_img in enumerate(augmented_images):
                    new_file = f"flipped_Img{i}_{file}"
                    save_path = os.path.join(augmentation_path, new_file)
                    augmented_img.save(save_path)

        QMessageBox.information(self, "Augmentation completed", "Flip augmentation completed!")


#Change Brightness        

    def change_brightness(self, image_directory, save_directory, num_images):
        datagen = ImageDataGenerator(brightness_range=[0.2, 2.1])
        self.augment_images(datagen, image_directory, save_directory, "Brightness_img", num_images)


    def change_brightness_images_with_imagau(self, image_directory, num_images, augmentation_path):
        seq = iaa.Multiply((0.4, 1.8))

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            img = Image.open(img_path)
            img_arr = np.array(img)

            # Apply augmentation multiple times
            for i in range(num_images):
                augmented_img_arr = seq.augment_image(img_arr)
                augmented_img = Image.fromarray(augmented_img_arr)

                # Save the augmented image
                new_file = f"Brightness_img{i}_{file}"
                save_path = os.path.join(augmentation_path, new_file)
                augmented_img.save(save_path)

        QMessageBox.information(self, "Augmentation completed", "Brightness augmentation completed!")

    def change_brightness_images_with_albumentations(self, image_directory, num_images, augmentation_path):
        # Create the augmentation transform
        transform = A.RandomBrightnessContrast(p=1.0)

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            img = Image.open(img_path)
            img_arr = np.array(img)

            # Apply augmentation multiple times
            for i in range(num_images):
                augmented = transform(image=img_arr)
                augmented_img_arr = augmented["image"]
                augmented_img = Image.fromarray(augmented_img_arr)

                # Save the augmented image
                new_file = f"Brightness_img{i}_{file}"
                save_path = os.path.join(augmentation_path, new_file)
                augmented_img.save(save_path)
        
        QMessageBox.information(self, "Augmentation completed", "Brightness augmentation completed!")


    def change_brightness_images_with_torchvision(self, image_directory, num_images, augmentation_path):
        # Create the augmentation transform
        transform = transforms.ColorJitter(brightness=0.5)

        # Iterate over each image in the directory
        image_files = os.listdir(image_directory)
        for file in image_files:
            img_path = os.path.join(image_directory, file)
            if os.path.isfile(img_path):
                # Load the image
                img = Image.open(img_path)

                # Apply augmentation to generate new images
                augmented_images = [transform(img) for _ in range(num_images)]

                # Save the augmented images
                for i, augmented_img in enumerate(augmented_images):
                    new_file = f"Brightness_img{i}_{file}"
                    save_path = os.path.join(augmentation_path, new_file)
                    augmented_img.save(save_path)

        QMessageBox.information(self, "Augmentation completed", "Brightness augmentation completed!")

#Add noise
    def add_salt_pepper_noise(self, image_directory, num_images, augmentation_path):
            for filename in os.listdir(image_directory):
                image_path = os.path.join(image_directory, filename)
                if os.path.isfile(image_path):
                    img = io.imread(image_path)
                    for i in range(num_images):
                        noisy_img = self.add_noise(img)
                        noisy_img = Image.fromarray(noisy_img)
                        new_file = f"Noisy{i}_{filename}"
                        save_path = os.path.join(augmentation_path, new_file)
                        noisy_img.save(save_path)
            QMessageBox.information(self, "Augmentation completed", "Noise Injected successfully!")
    def add_salt_pepper_noise_with_imagau(self, image_directory, num_images, augmentation_path):  
            print("hi")      
            for filename in os.listdir(image_directory):
                image_path = os.path.join(image_directory, filename)
                if os.path.isfile(image_path):
                    img = io.imread(image_path)
                    for i in range(num_images):
                        noisy_img = self.add_noise_with_imgaug(img)
                        noisy_img = Image.fromarray(noisy_img)
                        new_file = f"Noisy{i}_{filename}"
                        save_path = os.path.join(augmentation_path, new_file)
                        noisy_img.save(save_path)
            QMessageBox.information(self, "Augmentation completed", "Noise Injected successfully!")
    def add_salt_pepper_noise_with_albumentations(self, image_directory, num_images, augmentation_path):                  
            for filename in os.listdir(image_directory):
                image_path = os.path.join(image_directory, filename)
                if os.path.isfile(image_path):
                    img = io.imread(image_path)
                    for i in range(num_images):
                        noisy_img = self.add_noise_with_albumentations(img)
                        noisy_img = Image.fromarray(noisy_img)
                        new_file = f"Noisy{i}_{filename}"
                        save_path = os.path.join(augmentation_path, new_file)
                        noisy_img.save(save_path)
            QMessageBox.information(self, "Augmentation completed", "Noise Injected successfully!")
        
    def add_salt_pepper_noise_images_with_torchvision(self):
            QMessageBox.warning(self, "Warning", "Noise injection is not available with torchvision!")




    def augment_images(self, datagen, image_directory, save_directory, prefix, num_images):
        try:
            for filename in os.listdir(image_directory):
                image_path = os.path.join(image_directory, filename)
                if os.path.isfile(image_path):
                    img = io.imread(image_path)
                    img = img.reshape((1,) + img.shape)
                    save_prefix = os.path.join(save_directory, prefix)
                    self.save_augmented_images(datagen, img, save_prefix, num_images)
            QMessageBox.information(self,"Success", "Augmentation applied successfully.")
        except Exception as e:
            QMessageBox.warning(self,"Error", str(e))

    def save_augmented_images(self, datagen, img, save_prefix, num_images, save_format="png", batch_size=1):
        try:
            i = 0
            for batch in datagen.flow(
                img, batch_size=batch_size, save_to_dir=save_prefix, save_prefix=save_prefix, save_format=save_format
            ):
                i += 1
                if i >= num_images:
                    break
        except Exception as e:
            self.show_popup("Error", str(e))

    def show_popup(self, title, message):
        msg_box = QMessageBox()
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        msg_box.exec_()
        
if __name__ == "__main__":
    app = QApplication([])
    font = QFont()
    font.setPointSize(12)
    app.setFont(font)
    window = ImageAugmentor()
    window.show()
    app.exec_()
